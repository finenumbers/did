"""Tests for async XLSX export jobs and snapshots."""

from __future__ import annotations

import json
import threading
import time
from pathlib import Path

import pytest
import xlsxwriter
from openpyxl import load_workbook

from app.models.enums import InventoryKind
from app.services import numbers_export_jobs as jobs
from app.services.numbers_export import export_columns_schema, is_default_export_query


def test_is_default_export_query():
    assert is_default_export_query(
        filters=None, number_local_q=None, sort_by="abc_code", sort_dir="asc"
    )
    assert not is_default_export_query(
        filters={"region_name": ["x"]},
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    assert not is_default_export_query(
        filters=None, number_local_q="123", sort_by="abc_code", sort_dir="asc"
    )


def test_export_job_lifecycle(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("NUMBERS_EXPORT_DIR", str(tmp_path))
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setattr(jobs, "exports_dir", lambda: tmp_path)

    def fake_run_job(job_id: str) -> None:
        with jobs._lock:
            job = jobs._jobs.get(job_id)
            if job is None:
                jobs._active_builds = max(0, jobs._active_builds - 1)
                return
            job.status = "running"
            job.rows_total = 1
        out = tmp_path / f"job_{job_id}.xlsx"
        wb = xlsxwriter.Workbook(str(out))
        ws = wb.add_worksheet("Свободные")
        ws.write(0, 0, "Провайдер")
        ws.write(1, 0, "runexis")
        wb.close()
        with jobs._lock:
            job = jobs._jobs.get(job_id)
            if job is not None:
                job.status = "ready"
                job.path = str(out)
                job.rows_done = 1
                job.rows_total = 1
                job.finished_at = time.time()
            jobs._active_builds = max(0, jobs._active_builds - 1)

    monkeypatch.setattr(jobs, "_run_job", fake_run_job)

    job = jobs.create_export_job(
        inventory_kind=InventoryKind.free,
        filters={"abc_code": ["301"]},
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    for _ in range(50):
        j = jobs.get_job(job.id)
        assert j is not None
        if j.status in {"ready", "failed"}:
            break
        time.sleep(0.05)
    j = jobs.get_job(job.id)
    assert j is not None
    assert j.status == "ready"
    assert j.path and Path(j.path).is_file()

    wb = load_workbook(j.path, read_only=True)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[0][0] == "Провайдер"
        assert rows[1][0] == "runexis"
    finally:
        wb.close()

    get_settings.cache_clear()


def test_snapshot_fast_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("NUMBERS_EXPORT_DIR", str(tmp_path))
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setattr(jobs, "exports_dir", lambda: tmp_path)

    xlsx_path = tmp_path / "free_latest.xlsx"
    meta_path = tmp_path / "free_latest.meta.json"
    wb = xlsxwriter.Workbook(str(xlsx_path))
    ws = wb.add_worksheet("Свободные")
    ws.write(0, 0, "Провайдер")
    wb.close()
    meta_path.write_text(
        '{"count": 10, "max_last_seen_at": "2026-01-01T00:00:00+00:00",'
        ' "max_updated_at": "2026-01-01T00:00:00+00:00",'
        ' "sort_by": "abc_code", "sort_dir": "asc"}',
        encoding="utf-8",
    )

    monkeypatch.setattr(
        jobs,
        "catalog_fingerprint",
        lambda db, kind: {
            "count": 10,
            "max_last_seen_at": "2026-01-01T00:00:00+00:00",
            "max_updated_at": "2026-01-01T00:00:00+00:00",
        },
    )

    called = {"n": 0}

    def boom(*args, **kwargs):
        called["n"] += 1
        raise AssertionError("should use snapshot")

    monkeypatch.setattr(jobs, "_run_job", boom)
    monkeypatch.setattr(jobs, "_run_snapshot_write", boom)

    job = jobs.create_export_job(
        inventory_kind=InventoryKind.free,
        filters=None,
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    assert job.status == "ready"
    assert job.from_snapshot is True
    assert job.ticket
    assert job.phase == "ready"
    assert job.rows_total == 10
    assert called["n"] == 0
    assert job.path == str(xlsx_path)
    assert "ticket" in job.to_public()

    get_settings.cache_clear()


def _reset_export_state(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("NUMBERS_EXPORT_DIR", str(tmp_path))
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setattr(jobs, "exports_dir", lambda: tmp_path)
    with jobs._lock:
        jobs._jobs.clear()
        jobs._snapshot_writers.clear()
        jobs._active_builds = 0


def test_snapshot_stale_when_updated_at_changes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    _reset_export_state(tmp_path, monkeypatch)
    xlsx_path = tmp_path / "free_latest.xlsx"
    meta_path = tmp_path / "free_latest.meta.json"
    wb = xlsxwriter.Workbook(str(xlsx_path))
    ws = wb.add_worksheet("Свободные")
    ws.write(0, 0, "Провайдер")
    wb.close()
    meta_path.write_text(
        '{"count": 10, "max_last_seen_at": "2026-01-01T00:00:00+00:00",'
        ' "max_updated_at": "2026-01-01T00:00:00+00:00",'
        ' "sort_by": "abc_code", "sort_dir": "asc"}',
        encoding="utf-8",
    )
    fp = {
        "count": 10,
        "max_last_seen_at": "2026-01-01T00:00:00+00:00",
        "max_updated_at": "2026-08-19T12:00:00+00:00",
    }
    assert jobs.snapshot_is_fresh(InventoryKind.free, fp) is False


def test_snapshot_stale_without_or_when_columns_schema_changes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    _reset_export_state(tmp_path, monkeypatch)
    xlsx_path = tmp_path / "free_latest.xlsx"
    meta_path = tmp_path / "free_latest.meta.json"
    wb = xlsxwriter.Workbook(str(xlsx_path))
    ws = wb.add_worksheet("Свободные")
    ws.write(0, 0, "Провайдер")
    wb.close()
    schema = export_columns_schema()
    assert "Покупка (Входящая)" in schema
    assert "Покупка|" in schema or schema.endswith("Покупка")
    base_fp = {
        "count": 10,
        "max_last_seen_at": "2026-01-01T00:00:00+00:00",
        "max_updated_at": "2026-01-01T00:00:00+00:00",
        "columns_schema": schema,
    }
    meta_path.write_text(
        '{"count": 10, "max_last_seen_at": "2026-01-01T00:00:00+00:00",'
        ' "max_updated_at": "2026-01-01T00:00:00+00:00",'
        ' "sort_by": "abc_code", "sort_dir": "asc"}',
        encoding="utf-8",
    )
    assert jobs.snapshot_is_fresh(InventoryKind.free, base_fp) is False
    meta_path.write_text(
        '{"count": 10, "max_last_seen_at": "2026-01-01T00:00:00+00:00",'
        ' "max_updated_at": "2026-01-01T00:00:00+00:00",'
        ' "columns_schema": "old-headers",'
        ' "sort_by": "abc_code", "sort_dir": "asc"}',
        encoding="utf-8",
    )
    assert jobs.snapshot_is_fresh(InventoryKind.free, base_fp) is False
    meta_path.write_text(
        json.dumps(
            {
                "count": 10,
                "max_last_seen_at": "2026-01-01T00:00:00+00:00",
                "max_updated_at": "2026-01-01T00:00:00+00:00",
                "columns_schema": schema,
                "sort_by": "abc_code",
                "sort_dir": "asc",
            }
        ),
        encoding="utf-8",
    )
    assert jobs.snapshot_is_fresh(InventoryKind.free, base_fp) is True


def test_unfiltered_waits_on_snapshot_writer(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    _reset_export_state(tmp_path, monkeypatch)

    gate = threading.Event()
    started = threading.Event()
    calls = {"n": 0}

    def fake_export(*, path, on_progress=None, rows_total=None, **kwargs):
        calls["n"] += 1
        started.set()
        if on_progress:
            on_progress(5, 10, "writing")
        assert gate.wait(timeout=5)
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb = xlsxwriter.Workbook(str(path))
        ws = wb.add_worksheet("Свободные")
        ws.write(0, 0, "Провайдер")
        wb.close()
        if on_progress:
            on_progress(10, 10, "closing")
        return 10

    monkeypatch.setattr(jobs, "export_xlsx_job", fake_export)
    monkeypatch.setattr(
        jobs,
        "catalog_fingerprint",
        lambda db, kind: {
            "count": 10,
            "max_last_seen_at": "2026-01-01T00:00:00+00:00",
            "max_updated_at": "2026-01-01T00:00:00+00:00",
        },
    )

    job1 = jobs.create_export_job(
        inventory_kind=InventoryKind.free,
        filters=None,
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    assert started.wait(timeout=2)
    assert job1.status in {"queued", "running"}

    job2 = jobs.create_export_job(
        inventory_kind=InventoryKind.free,
        filters=None,
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    assert job2.status in {"queued", "running"}
    assert calls["n"] == 1

    gate.set()
    for jid in (job1.id, job2.id):
        for _ in range(100):
            j = jobs.get_job(jid)
            assert j is not None
            if j.status in {"ready", "failed"}:
                break
            time.sleep(0.05)
        j = jobs.get_job(jid)
        assert j is not None
        assert j.status == "ready"
        assert j.from_snapshot is True
        assert j.ticket
        assert j.phase == "ready"
        assert j.path and Path(j.path).name == "free_latest.xlsx"
    assert calls["n"] == 1

    from app.core.config import get_settings

    get_settings.cache_clear()


def test_free_fast_path_while_purchased_snapshot_writes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    _reset_export_state(tmp_path, monkeypatch)
    xlsx_path = tmp_path / "free_latest.xlsx"
    meta_path = tmp_path / "free_latest.meta.json"
    wb = xlsxwriter.Workbook(str(xlsx_path))
    ws = wb.add_worksheet("Свободные")
    ws.write(0, 0, "Провайдер")
    wb.close()
    meta_path.write_text(
        '{"count": 10, "max_last_seen_at": "2026-01-01T00:00:00+00:00",'
        ' "max_updated_at": "2026-01-01T00:00:00+00:00",'
        ' "sort_by": "abc_code", "sort_dir": "asc"}',
        encoding="utf-8",
    )
    monkeypatch.setattr(
        jobs,
        "catalog_fingerprint",
        lambda db, kind: {
            "count": 10,
            "max_last_seen_at": "2026-01-01T00:00:00+00:00",
            "max_updated_at": "2026-01-01T00:00:00+00:00",
        },
    )

    def boom(*args, **kwargs):
        raise AssertionError("should use free snapshot")

    monkeypatch.setattr(jobs, "_run_job", boom)
    monkeypatch.setattr(jobs, "_run_snapshot_write", boom)

    writer = jobs._KindWriter(kind=InventoryKind.purchased)
    with jobs._lock:
        jobs._snapshot_writers[InventoryKind.purchased] = writer
        jobs._active_builds = 1

    job = jobs.create_export_job(
        inventory_kind=InventoryKind.free,
        filters=None,
        number_local_q=None,
        sort_by="abc_code",
        sort_dir="asc",
    )
    assert job.status == "ready"
    assert job.from_snapshot is True
    writer.done.set()
    with jobs._lock:
        jobs._snapshot_writers.pop(InventoryKind.purchased, None)
        jobs._active_builds = 0

    from app.core.config import get_settings

    get_settings.cache_clear()


def test_job_ticket_matches():
    job = jobs.ExportJob(
        id="abc",
        inventory_kind=InventoryKind.free,
        status="ready",
        ticket="a" * 32,
    )
    assert jobs.job_ticket_matches(job, "a" * 32)
    assert not jobs.job_ticket_matches(job, "b" * 32)
    assert not jobs.job_ticket_matches(job, None)
    assert "ticket" in job.to_public()
    running = jobs.ExportJob(
        id="def",
        inventory_kind=InventoryKind.free,
        status="running",
        ticket="a" * 32,
    )
    assert "ticket" not in running.to_public()


def test_download_with_ticket(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    _reset_export_state(tmp_path, monkeypatch)
    from fastapi import HTTPException

    from app.api.routes.numbers import download_export_job

    xlsx_path = tmp_path / "free_latest.xlsx"
    wb = xlsxwriter.Workbook(str(xlsx_path))
    ws = wb.add_worksheet("Свободные")
    ws.write(0, 0, "Провайдер")
    wb.close()
    job = jobs.ExportJob(
        id="ticketjob1",
        inventory_kind=InventoryKind.free,
        status="ready",
        phase="ready",
        filename="free_numbers.xlsx",
        path=str(xlsx_path),
        ticket="t" * 32,
        from_snapshot=True,
    )
    with jobs._lock:
        jobs._jobs[job.id] = job

    resp = download_export_job(job.id, ticket=job.ticket)
    assert resp.filename == "free_numbers.xlsx"

    with pytest.raises(HTTPException) as exc:
        download_export_job(job.id, ticket="x" * 32)
    assert exc.value.status_code == 404


def test_export_download_ticket_bypasses_middleware(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("ADMIN_USERNAME", "admin")
    monkeypatch.setenv("ADMIN_PASSWORD", "secret")
    from app.core.config import get_settings

    get_settings.cache_clear()
    from starlette.applications import Starlette
    from starlette.responses import PlainTextResponse
    from starlette.routing import Route
    from starlette.testclient import TestClient

    from app.api.auth import AdminAuthMiddleware

    async def dummy(_request):
        return PlainTextResponse("ok")

    app = Starlette(
        routes=[
            Route("/api/v1/numbers/export-jobs/abc/download", dummy),
            Route("/api/v1/numbers/export-jobs/abc", dummy),
        ]
    )
    app.add_middleware(AdminAuthMiddleware)
    client = TestClient(app)
    try:
        denied = client.get("/api/v1/numbers/export-jobs/abc/download")
        assert denied.status_code == 401
        allowed = client.get("/api/v1/numbers/export-jobs/abc/download?ticket=secret")
        assert allowed.status_code == 200
        status_denied = client.get("/api/v1/numbers/export-jobs/abc")
        assert status_denied.status_code == 401
    finally:
        monkeypatch.delenv("ADMIN_USERNAME", raising=False)
        monkeypatch.delenv("ADMIN_PASSWORD", raising=False)
        get_settings.cache_clear()
