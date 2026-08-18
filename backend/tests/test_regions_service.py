from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.services.regions_service import (
    REGIONS_XLSX_HEADERS,
    RegionsService,
    parse_regions_xlsx,
)
from app.services.xlsx_style import StyledSheetWriter, open_styled_workbook


def _db_with_rows(rows: list) -> MagicMock:
    db = MagicMock()
    db.scalars.return_value.all.return_value = rows
    return db


def _xlsx_bytes(path: Path, rows: list[list[object]]) -> bytes:
    wb = open_styled_workbook(str(path), constant_memory=True)
    try:
        ws = wb.add_worksheet("Sheet1")
        writer = StyledSheetWriter(wb, ws, REGIONS_XLSX_HEADERS)
        for row in rows:
            writer.write_row(row)
        writer.finalize()
    finally:
        wb.close()
    return path.read_bytes()


def test_list_cities_includes_abc_and_same_geo():
    id_495 = uuid4()
    id_499 = uuid4()
    rows = [
        SimpleNamespace(
            id=id_495,
            abc="495",
            digit_capacity=7,
            city_name="Москва",
            region_name="Москва",
        ),
        SimpleNamespace(
            id=id_499,
            abc="499",
            digit_capacity=7,
            city_name="Москва",
            region_name="Москва",
        ),
    ]
    items = RegionsService(_db_with_rows(rows)).list_cities()
    assert [(i.abc, i.digit_capacity, i.city_name, i.region_name) for i in items] == [
        ("495", 7, "Москва", "Москва"),
        ("499", 7, "Москва", "Москва"),
    ]


def test_parse_numeric_excel_cells(tmp_path: Path):
    data = _xlsx_bytes(
        tmp_path / "num.xlsx",
        [[495, 7.0, "Москва", "Москва"]],
    )
    rows = parse_regions_xlsx(data)
    assert len(rows) == 1
    assert rows[0].abc == "495"
    assert rows[0].digit_capacity == 7
    assert rows[0].city_name == "Москва"
    assert rows[0].region_name == "Москва"


def test_parse_two_abc_same_city_region(tmp_path: Path):
    data = _xlsx_bytes(
        tmp_path / "msk.xlsx",
        [
            ["495", 7, "Москва", "Москва"],
            ["499", 7, "Москва", "Москва"],
        ],
    )
    rows = parse_regions_xlsx(data)
    assert [r.abc for r in rows] == ["495", "499"]
    assert all(r.city_name == "Москва" and r.region_name == "Москва" for r in rows)


def test_parse_header_only_is_empty(tmp_path: Path):
    assert parse_regions_xlsx(_xlsx_bytes(tmp_path / "empty.xlsx", [])) == []


def test_parse_rejects_duplicate_abc(tmp_path: Path):
    data = _xlsx_bytes(
        tmp_path / "dup.xlsx",
        [
            ["495", 7, "Москва", "Москва"],
            ["495", 7, "Химки", "МО"],
        ],
    )
    with pytest.raises(ValueError, match="повторяется ABC"):
        parse_regions_xlsx(data)


def test_parse_rejects_capacity_and_abc_len(tmp_path: Path):
    with pytest.raises(ValueError, match="разрядность"):
        parse_regions_xlsx(
            _xlsx_bytes(tmp_path / "cap.xlsx", [["495", 4, "Москва", "Москва"]])
        )
    with pytest.raises(ValueError, match="давать 10"):
        parse_regions_xlsx(
            _xlsx_bytes(tmp_path / "len.xlsx", [["495", 6, "Москва", "Москва"]])
        )


def test_parse_rejects_wrong_headers(tmp_path: Path):
    path = tmp_path / "bad.xlsx"
    wb = open_styled_workbook(str(path), constant_memory=True)
    try:
        ws = wb.add_worksheet("data")
        writer = StyledSheetWriter(wb, ws, ["A", "B", "C", "D"])
        writer.write_row(["495", 7, "Москва", "Москва"])
        writer.finalize()
    finally:
        wb.close()
    with pytest.raises(ValueError, match="Заголовки"):
        parse_regions_xlsx(path.read_bytes())


def test_replace_skips_db_on_parse_error():
    db = MagicMock()
    with pytest.raises(ValueError):
        RegionsService(db).replace_from_xlsx(b"not-xlsx")
    db.execute.assert_not_called()
    db.commit.assert_not_called()
    db.add_all.assert_not_called()


def test_replace_from_xlsx_writes_parsed_rows(tmp_path: Path):
    db = MagicMock()
    data = _xlsx_bytes(
        tmp_path / "ok.xlsx",
        [
            ["3842", 6, "Кемерово", "Кемеровская область"],
            ["495", 7, "Москва", "Москва"],
        ],
    )
    out = RegionsService(db).replace_from_xlsx(data)
    assert out.ok is True
    assert out.count == 2
    db.execute.assert_called_once()
    added = db.add_all.call_args[0][0]
    assert {(row.abc, row.digit_capacity, row.city_name) for row in added} == {
        ("3842", 6, "Кемерово"),
        ("495", 7, "Москва"),
    }
    db.commit.assert_called_once()


def test_write_xlsx_columns_and_roundtrip(tmp_path: Path):
    id_495 = uuid4()
    rows = [
        SimpleNamespace(
            id=id_495,
            abc="495",
            digit_capacity=7,
            city_name="Москва",
            region_name="Москва",
        )
    ]
    path = tmp_path / "out.xlsx"
    count = RegionsService(_db_with_rows(rows)).write_xlsx(path)
    assert count == 1
    wb = load_workbook(path)
    try:
        sheet = wb.active
        assert [sheet.cell(1, col).value for col in range(1, 5)] == list(REGIONS_XLSX_HEADERS)
        assert sheet.cell(2, 1).value in ("495", 495)
        assert sheet.cell(2, 2).value == 7
        assert sheet.cell(2, 3).value == "Москва"
        assert sheet.cell(2, 4).value == "Москва"
    finally:
        wb.close()
    parsed = parse_regions_xlsx(path.read_bytes())
    assert parsed[0].abc == "495"
    assert parsed[0].digit_capacity == 7
