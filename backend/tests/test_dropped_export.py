"""Tests for sync dropped-numbers XLSX collector."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.models.enums import InventoryKind, MappingConfidence
from app.modules.sync_engine.dropped_export import (
    begin_dropped_export,
    end_dropped_export,
    record_number_drops,
    split_dedupe_drops,
    write_dropped_xlsx,
)
from app.providers.dto.numbers import NormalizedNumber


def _num(key: str, raw: dict | None = None) -> NormalizedNumber:
    return NormalizedNumber(
        inventory_kind=InventoryKind.free,
        provider_number_key=key,
        msisdn=key,
        city_external_id=None,
        region_external_id=None,
        city_name=None,
        region_name=None,
        buy_price=None,
        period_price=None,
        status_raw=None,
        mapping_confidence=MappingConfidence.medium,
        normalized_payload={},
        raw_payload=raw or {"phone": key},
    )


def _patch_xlsx_path(monkeypatch, path: Path) -> None:
    monkeypatch.setenv("SYNC_DROPPED_XLSX_PATH", str(path))
    from app.core.config import get_settings

    get_settings.cache_clear()
    monkeypatch.setattr(
        "app.modules.sync_engine.dropped_export.get_settings",
        lambda: type("S", (), {"sync_dropped_xlsx_path": str(path)})(),
    )


def test_split_dedupe_drops_last_wins():
    nums = [
        _num("79001111111", {"n": 1}),
        _num("79002222222", {"n": 2}),
        _num("79001111111", {"n": 3}),
    ]
    dropped, kept = split_dedupe_drops(nums)
    assert len(kept) == 2
    assert len(dropped) == 1
    assert dropped[0].raw_payload == {"n": 1}
    kept_by_key = {n.provider_number_key: n for n in kept}
    assert kept_by_key["79001111111"].raw_payload == {"n": 3}


def test_write_dropped_xlsx(tmp_path: Path, monkeypatch):
    path = tmp_path / "sync_dropped_latest.xlsx"
    _patch_xlsx_path(monkeypatch, path)

    begin_dropped_export()
    try:
        record_number_drops(
            provider="uis",
            inventory_kind="free",
            unmapped_raw=[{"phone_number": ""}],
            numbers=[
                _num("79001111111", {"n": 1}),
                _num("79001111111", {"n": 2}),
            ],
        )
        meta = write_dropped_xlsx()
    finally:
        end_dropped_export()

    assert meta["available"] is True
    assert meta["unmapped"] == 1
    assert meta["duplicates"] == 1
    assert path.is_file()

    wb = load_workbook(path)
    assert "unmapped" in wb.sheetnames
    assert "duplicates" in wb.sheetnames
    assert wb["unmapped"].max_row == 2
    assert wb["duplicates"].max_row == 2


def test_begin_preserves_previous_file(tmp_path: Path, monkeypatch):
    path = tmp_path / "sync_dropped_latest.xlsx"
    path.write_bytes(b"old")
    _patch_xlsx_path(monkeypatch, path)
    begin_dropped_export()
    try:
        assert path.read_bytes() == b"old"
        meta = write_dropped_xlsx()  # empty collector → preserve attempt
        assert meta.get("preserved_previous") is True
        assert meta["available"] is False  # unreadable garbage bytes
        assert path.read_bytes() == b"old"
    finally:
        end_dropped_export()


def test_preserve_reads_honest_counts_from_xlsx(tmp_path: Path, monkeypatch):
    path = tmp_path / "sync_dropped_latest.xlsx"
    wb = Workbook()
    # openpyxl creates a default sheet; replace with our sheets
    default = wb.active
    wb.remove(default)
    ws_u = wb.create_sheet("unmapped")
    ws_d = wb.create_sheet("duplicates")
    headers = ["provider", "inventory_kind", "provider_number_key", "outcome", "raw_payload"]
    ws_u.append(headers)
    ws_d.append(headers)
    for i in range(3):
        ws_u.append(["uis", "free", f"u{i}", "dropped", "{}"])
    for i in range(2):
        ws_d.append(["uis", "free", f"d{i}", "dropped", "{}"])
    wb.save(path)

    _patch_xlsx_path(monkeypatch, path)
    begin_dropped_export()
    try:
        meta = write_dropped_xlsx()
    finally:
        end_dropped_export()

    assert meta["available"] is True
    assert meta["preserved_previous"] is True
    assert meta["unmapped"] == 3
    assert meta["duplicates"] == 2
    assert meta["generated_at"]
    assert path.is_file()
