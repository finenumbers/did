"""Shared XLSX export styling (Calibri 11, no fills, autofit, autofilter)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.xlsx_style import (
    FONT_NAME,
    FONT_SIZE,
    StyledSheetWriter,
    excel_col_width,
    open_styled_workbook,
)


def test_styled_sheet_writer_format(tmp_path: Path):
    path = tmp_path / "styled.xlsx"
    wb = open_styled_workbook(str(path), constant_memory=True)
    try:
        ws = wb.add_worksheet("data")
        writer = StyledSheetWriter(wb, ws, ["Name", "City"])
        writer.write_row(["Alice", "Moscow"])
        writer.write_row(["Bob", "Saint Petersburg"])
        writer.finalize()
    finally:
        wb.close()

    loaded = load_workbook(path)
    try:
        sheet = loaded.active
        assert sheet["A1"].value == "Name"
        assert sheet["A1"].font.bold is True
        assert sheet["A1"].font.name == FONT_NAME
        assert sheet["A1"].font.size == FONT_SIZE
        assert sheet["A1"].alignment.horizontal == "center"
        assert sheet["A1"].border.left.style == "thin"
        assert sheet["B2"].value == "Moscow"
        assert sheet["B2"].font.name == FONT_NAME
        assert sheet["B2"].font.size == FONT_SIZE
        assert sheet["B2"].border.left.style == "thin"
        # No colored pattern fills on header/data cells.
        for cell in (sheet["A1"], sheet["B1"], sheet["A2"], sheet["B2"]):
            fill = cell.fill
            assert getattr(fill, "patternType", None) in (None, "none")
            assert getattr(fill, "fgColor", None) is None or fill.fgColor.rgb in (
                None,
                "00000000",
                "0",
            )
        assert sheet.auto_filter.ref == "A1:B3"
        # Wider of header "City"(4) vs "Saint Petersburg"(16) → padded width.
        expected = excel_col_width(len("Saint Petersburg"))
        assert sheet.column_dimensions["B"].width >= expected - 1
        assert sheet.column_dimensions["B"].width <= expected + 2
    finally:
        loaded.close()


def test_styled_sheet_writer_preset_widths_skip_content(tmp_path: Path):
    path = tmp_path / "preset.xlsx"
    wb = open_styled_workbook(str(path), constant_memory=True)
    try:
        ws = wb.add_worksheet("data")
        writer = StyledSheetWriter(
            wb, ws, ["City"], track_content_width=False, min_chars=[12]
        )
        writer.write_row(["Saint Petersburg"])
        writer.finalize()
    finally:
        wb.close()

    loaded = load_workbook(path)
    try:
        sheet = loaded.active
        expected = excel_col_width(12)
        assert sheet.column_dimensions["A"].width >= expected - 1
        assert sheet.column_dimensions["A"].width <= expected + 2
        assert sheet.auto_filter.ref == "A1:A2"
    finally:
        loaded.close()


def test_styled_sheet_writer_fill_and_content_width(tmp_path: Path):
    path = tmp_path / "filled.xlsx"
    wb = open_styled_workbook(str(path), constant_memory=True)
    try:
        ws = wb.add_worksheet("data")
        writer = StyledSheetWriter(wb, ws, ["City"], min_chars=[8])
        writer.write_row(["Saint Petersburg"], fill_color="#C6EFCE")
        writer.finalize()
    finally:
        wb.close()

    loaded = load_workbook(path)
    try:
        sheet = loaded.active
        fill = sheet["A2"].fill
        assert getattr(fill, "patternType", None) == "solid"
        expected = excel_col_width(len("Saint Petersburg"))
        assert sheet.column_dimensions["A"].width >= expected - 1
        assert sheet["A2"].border.left.style == "thin"
    finally:
        loaded.close()
