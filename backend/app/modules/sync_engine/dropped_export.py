"""Collect unmapped/duplicate sync rows and write latest-only XLSX report."""

from __future__ import annotations

import json
import logging
import os
import threading
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.config import get_settings
from app.providers.dto.numbers import NormalizedNumber

logger = logging.getLogger(__name__)

_lock = threading.Lock()
_collector: DroppedCollector | None = None


@dataclass
class DroppedRow:
    provider: str
    inventory_kind: str
    reason: str  # unmapped | duplicate
    provider_number_key: str | None
    raw_payload: dict[str, Any]


@dataclass
class DroppedCollector:
    rows: list[DroppedRow] = field(default_factory=list)

    def add_unmapped(
        self, provider: str, inventory_kind: str, raw_payload: dict[str, Any]
    ) -> None:
        self.rows.append(
            DroppedRow(
                provider=provider,
                inventory_kind=inventory_kind,
                reason="unmapped",
                provider_number_key=None,
                raw_payload=raw_payload if isinstance(raw_payload, dict) else {},
            )
        )

    def add_duplicate(
        self,
        provider: str,
        inventory_kind: str,
        *,
        provider_number_key: str | None,
        raw_payload: dict[str, Any],
    ) -> None:
        self.rows.append(
            DroppedRow(
                provider=provider,
                inventory_kind=inventory_kind,
                reason="duplicate",
                provider_number_key=provider_number_key,
                raw_payload=raw_payload if isinstance(raw_payload, dict) else {},
            )
        )

    def counts(self) -> tuple[int, int]:
        unmapped = sum(1 for r in self.rows if r.reason == "unmapped")
        duplicates = sum(1 for r in self.rows if r.reason == "duplicate")
        return unmapped, duplicates


def dropped_xlsx_path() -> Path:
    return Path(get_settings().sync_dropped_xlsx_path)


def get_collector() -> DroppedCollector | None:
    return _collector


def begin_dropped_export() -> DroppedCollector:
    """Start a new unified-run collector; keep previous XLSX until a successful write."""
    global _collector
    with _lock:
        _collector = DroppedCollector()
        path = dropped_xlsx_path()
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            logger.exception("Failed to ensure dropped export directory %s", path.parent)
        return _collector


def end_dropped_export() -> None:
    global _collector
    with _lock:
        _collector = None


def split_dedupe_drops(
    numbers: list[NormalizedNumber],
) -> tuple[list[NormalizedNumber], list[NormalizedNumber]]:
    """Mirror persist dedupe (last wins). Return (dropped_duplicates, kept)."""
    kept_map: dict[str, NormalizedNumber] = {}
    dropped: list[NormalizedNumber] = []
    for num in numbers:
        key = num.provider_number_key
        if not key:
            continue
        if key in kept_map:
            dropped.append(kept_map[key])
        kept_map[key] = num
    return dropped, list(kept_map.values())


def record_number_drops(
    *,
    provider: str,
    inventory_kind: str,
    unmapped_raw: list[dict[str, Any]] | None,
    numbers: list[NormalizedNumber],
) -> None:
    collector = get_collector()
    if collector is None:
        return
    for raw in unmapped_raw or []:
        if isinstance(raw, dict):
            collector.add_unmapped(provider, inventory_kind, raw)
    dropped, _kept = split_dedupe_drops(numbers)
    for num in dropped:
        collector.add_duplicate(
            provider,
            inventory_kind,
            provider_number_key=num.provider_number_key,
            raw_payload=num.raw_payload or {},
        )


def _counts_from_existing_xlsx(path: Path) -> tuple[int, int] | None:
    """Return (unmapped, duplicates) data rows from existing report, or None if unreadable."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            unmapped_n = 0
            duplicates_n = 0
            if "unmapped" in wb.sheetnames:
                # subtract header row when present
                unmapped_n = max(0, (wb["unmapped"].max_row or 0) - 1)
            if "duplicates" in wb.sheetnames:
                duplicates_n = max(0, (wb["duplicates"].max_row or 0) - 1)
            return unmapped_n, duplicates_n
        finally:
            wb.close()
    except Exception:
        logger.exception("Failed to read preserved dropped export %s", path)
        return None


def write_dropped_xlsx() -> dict[str, Any]:
    """Write collector rows atomically; preserve previous file when collector is empty."""
    collector = get_collector() or DroppedCollector()
    path = dropped_xlsx_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    if not collector.rows:
        if not dropped_xlsx_exists():
            return {
                "available": False,
                "path_basename": None,
                "unmapped": 0,
                "duplicates": 0,
                "preserved_previous": False,
                "generated_at": None,
            }
        counts = _counts_from_existing_xlsx(path)
        if counts is None:
            return {
                "available": False,
                "path_basename": path.name,
                "unmapped": 0,
                "duplicates": 0,
                "preserved_previous": True,
                "generated_at": None,
                "error": "preserved_unreadable",
            }
        unmapped_n, duplicates_n = counts
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=UTC).isoformat()
        return {
            "available": True,
            "path_basename": path.name,
            "unmapped": unmapped_n,
            "duplicates": duplicates_n,
            "preserved_previous": True,
            "generated_at": mtime,
        }

    wb = Workbook(write_only=True)
    ws_u = wb.create_sheet("unmapped")
    ws_d = wb.create_sheet("duplicates")

    headers = ["provider", "inventory_kind", "provider_number_key", "outcome", "raw_payload"]
    ws_u.append(headers)
    ws_d.append(headers)

    unmapped_n = 0
    duplicates_n = 0
    for row in collector.rows:
        payload = json.dumps(row.raw_payload, ensure_ascii=False, default=str)
        values = [
            row.provider,
            row.inventory_kind,
            row.provider_number_key or "",
            "dropped",
            payload,
        ]
        if row.reason == "unmapped":
            ws_u.append(values)
            unmapped_n += 1
        else:
            ws_d.append(values)
            duplicates_n += 1

    tmp_path = path.with_name(path.name + ".tmp")
    wb.save(tmp_path)
    os.replace(tmp_path, path)
    generated_at = datetime.now(UTC).isoformat()
    meta = {
        "available": True,
        "path_basename": path.name,
        "unmapped": unmapped_n,
        "duplicates": duplicates_n,
        "generated_at": generated_at,
    }
    logger.info(
        "Wrote sync dropped export %s unmapped=%s duplicates=%s",
        path,
        unmapped_n,
        duplicates_n,
    )
    return meta


def dropped_xlsx_exists() -> bool:
    path = dropped_xlsx_path()
    return path.is_file() and path.stat().st_size > 0
