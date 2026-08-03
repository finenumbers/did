"""XLSX export of filtered numbers catalog (Unicode / openpyxl)."""

from __future__ import annotations

from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.db import SessionLocal
from app.models.catalog import NumbersCatalogNormalized
from app.models.enums import InventoryKind
from app.services.numbers_service import NumbersService

# UI column order / Russian headers (matches NumbersTable)
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("provider_code", "Провайдер"),
    ("abc_code", "ABC"),
    ("number_category", "Категория"),
    ("number_local", "Номер"),
    ("status_raw", "Статус"),
    ("region_name", "Регион"),
    ("city_name", "Город"),
    ("buy_price", "Покупка"),
    ("period_price", "Абонплата"),
    ("mask", "Маска"),
    ("display_mask", "Display mask"),
    ("book_date", "Book date"),
    ("number_type", "Тип"),
    ("points", "Баллы"),
    ("date_from", "date_from"),
    ("last_operation_date", "last_operation_date"),
    ("operator", "Оператор"),
    ("operator_id", "operator_id"),
    ("manager_id", "manager_id"),
    ("notes", "notes"),
    ("abcdef", "abcdef"),
    ("order_id", "order_id"),
    ("doc_status", "doc_status"),
    ("doc_required", "doc_required"),
    ("order_doc_required", "order_doc_required"),
    ("sign", "sign"),
    ("tariff", "Тариф"),
    ("class", "Класс"),
    ("partner", "Партнёр"),
    ("project", "Проект"),
    ("equipment", "Оборудование"),
    ("mapping_confidence", "confidence"),
    ("last_seen_at", "Обновлено"),
]

COLUMN_WIDTHS: dict[str, float] = {
    "provider_code": 12,
    "abc_code": 8,
    "number_category": 16,
    "number_local": 12,
    "status_raw": 12,
    "region_name": 28,
    "city_name": 22,
    "buy_price": 12,
    "period_price": 12,
    "mask": 14,
    "display_mask": 16,
    "book_date": 14,
    "number_type": 10,
    "points": 10,
    "notes": 24,
    "tariff": 16,
    "class": 14,
    "operator": 16,
    "partner": 16,
    "project": 16,
    "equipment": 16,
    "mapping_confidence": 12,
    "last_seen_at": 20,
}

_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="D9E2EC")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BATCH = 5_000


def _format_price(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        n = int(Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError, TypeError):
        return str(value)
    sign = "-" if n < 0 else ""
    digits = str(abs(n))
    parts: list[str] = []
    while digits:
        parts.append(digits[-3:])
        digits = digits[:-3]
    return sign + " ".join(reversed(parts))


def _format_points(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{Decimal(str(value)):.2f}"
    except (InvalidOperation, ValueError, TypeError):
        return str(value)


def _cell_value(key: str, row: NumbersCatalogNormalized, provider_code: str) -> Any:
    if key == "provider_code":
        return provider_code
    if key == "class":
        return row.number_class or ""
    if key == "buy_price":
        return _format_price(row.buy_price)
    if key == "period_price":
        return _format_price(row.period_price)
    if key == "points":
        return _format_points(row.points)
    if key == "mapping_confidence":
        conf = row.mapping_confidence
        return conf.value if hasattr(conf, "value") else (str(conf) if conf else "")
    if key == "last_seen_at":
        ts = row.last_seen_at
        if isinstance(ts, datetime):
            return ts.isoformat()
        return str(ts) if ts else ""
    val = getattr(row, key, None)
    return "" if val is None else val


def _header_cell(ws: Any, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(ws, value=value)
    cell.border = _THIN
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _HEADER_ALIGN
    return cell


class NumbersExportService:
    def __init__(self, db: Session):
        self.db = db
        self.numbers = NumbersService(db)

    def export_xlsx(
        self,
        *,
        inventory_kind: InventoryKind,
        path: str | Path,
        filters: dict[str, list[str]] | None = None,
        number_local_q: str | None = None,
        sort_by: str | None = None,
        sort_dir: str = "asc",
    ) -> int:
        """Write filtered catalog to path. Returns data row count (excluding header)."""
        stmt = self.numbers._base_stmt(inventory_kind, is_currently_present=True)
        stmt = self.numbers.apply_catalog_filters(
            stmt,
            filters=filters,
            number_local_q=number_local_q,
        )
        stmt = stmt.order_by(*self.numbers.order_by_clauses(sort_by, sort_dir))

        sheet_title = "Свободные" if inventory_kind == InventoryKind.free else "Купленные"
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_title)

        for idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS.get(key, 14)

        # Header: bold + fill + borders. Data rows are plain values for speed
        # (bordering ~370k×33 cells made free export appear to hang in the browser).
        ws.append([_header_cell(ws, header) for _, header in EXPORT_COLUMNS])

        keys = [key for key, _ in EXPORT_COLUMNS]
        row_count = 0
        result = self.db.execute(stmt.execution_options(yield_per=_BATCH))
        for row, code in result:
            provider_code = code.value if hasattr(code, "value") else str(code)
            ws.append([_cell_value(key, row, provider_code) for key in keys])
            row_count += 1

        wb.save(path)
        wb.close()
        return row_count


def export_xlsx_job(
    *,
    inventory_kind: InventoryKind,
    path: str,
    filters: dict[str, list[str]] | None,
    number_local_q: str | None,
    sort_by: str | None,
    sort_dir: str,
) -> int:
    """Run export in a dedicated DB session (safe for asyncio.to_thread)."""
    db = SessionLocal()
    try:
        return NumbersExportService(db).export_xlsx(
            inventory_kind=inventory_kind,
            path=path,
            filters=filters,
            number_local_q=number_local_q,
            sort_by=sort_by,
            sort_dir=sort_dir,
        )
    finally:
        db.close()
