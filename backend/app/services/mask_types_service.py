from __future__ import annotations

import json
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Protocol
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import Select, cast, func, or_, select
from sqlalchemy.orm import Session
from sqlalchemy.sql import ColumnElement
from sqlalchemy.types import String

from app.models.mask_types import MaskType
from app.modules.catalog.apply_mask_types import normalize_key_part
from app.modules.catalog.beauty_mask import (
    all_beauty_masks,
    canonical_beauty_masks,
    mask_digit_capacity,
)
from app.modules.catalog.mask_type_policy import (
    is_required_key,
    mask_row_fill_color,
    normalize_import_category,
    required_categories,
)
from app.schemas.common import Page
from app.schemas.mask_types import MaskTypeItem, MaskTypesLoadResult
from app.schemas.numbers import FacetItem, FacetResponse
from app.services.xlsx_style import StyledSheetWriter, open_styled_workbook

MASK_TYPES_XLSX_HEADERS = (
    "Разрядность",
    "Категория",
    "ABC",
    "Маска",
    "Тип",
    "Премиум",
    "Покупка",
)
MASK_TYPES_XLSX_SHEET = "Маски"
MAX_IMPORT_BYTES = 5 * 1024 * 1024
_IMPORT_WIDTH = len(MASK_TYPES_XLSX_HEADERS)
EMPTY_TOKEN = "__empty__"


@dataclass(frozen=True)
class MaskTypeImportRow:
    digit_capacity: str
    category: str
    abc: str
    mask: str
    type_label: str | None
    premium: Decimal | None
    purchase: Decimal | None


class _KeyedRow(Protocol):
    digit_capacity: str
    category: str
    abc: str
    mask: str


def _cell_text(value: object) -> str:
    return normalize_key_part(value)


def _nullable_cell(value: object) -> str | None:
    text = _cell_text(value)
    return text or None


def _row_empty(values: list[object]) -> bool:
    return all(_cell_text(v) == "" for v in values)


def _parse_price_cell(value: object, *, row: int, column: str) -> Decimal | None:
    if isinstance(value, bool):
        raise ValueError(f"Строка {row}: некорректная цена в столбце {column}")
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    text = _cell_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Строка {row}: некорректная цена в столбце {column}") from exc


def _xlsx_price(value: Decimal | None) -> object:
    if value is None:
        return ""
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _format_price_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        n = int(Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError, TypeError):
        return str(value)
    sign = "-" if n < 0 else ""
    digits = str(abs(n))
    parts: list[str] = []
    while digits:
        parts.append(digits[-3:])
        digits = digits[:-3]
    return sign + " ".join(reversed(parts))


def _parse_price_token(token: str) -> Decimal | None:
    raw = token.replace(" ", "").replace(",", ".")
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def _item_from_row(row: MaskType) -> MaskTypeItem:
    return MaskTypeItem(
        id=row.id,
        digit_capacity=row.digit_capacity,
        category=row.category,
        abc=row.abc,
        mask=row.mask,
        type_label=row.type_label,
        premium=row.premium,
        purchase=row.purchase,
    )


def row_key(row: _KeyedRow) -> tuple[str, str, str, str]:
    return (row.digit_capacity or "", row.category or "", row.abc or "", row.mask)


def apply_payload(row: MaskType, item: MaskTypeImportRow) -> None:
    row.type_label = item.type_label
    row.premium = item.premium
    row.purchase = item.purchase


def parse_mask_types_xlsx(data: bytes) -> list[MaskTypeImportRow]:
    if not data:
        raise ValueError("Пустой файл")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("Файл слишком большой")
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать XLSX") from exc
    try:
        if not wb.worksheets:
            raise ValueError("В файле нет листов")
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ValueError("Нет заголовков") from exc
        header_cells = list(header or ())
        while len(header_cells) < _IMPORT_WIDTH:
            header_cells.append(None)
        got = tuple(_cell_text(c) for c in header_cells[:_IMPORT_WIDTH])
        if got != MASK_TYPES_XLSX_HEADERS:
            raise ValueError(
                "Заголовки должны быть: Разрядность, Категория, ABC, Маска, Тип, Премиум, Покупка"
            )
        canonical = canonical_beauty_masks()
        out: list[MaskTypeImportRow] = []
        seen: set[tuple[str, str, str, str]] = set()
        for idx, raw in enumerate(rows_iter, start=2):
            cells = list(raw or ())
            while len(cells) < _IMPORT_WIDTH:
                cells.append(None)
            chunk = cells[:_IMPORT_WIDTH]
            if _row_empty(chunk):
                continue
            category = _cell_text(chunk[1])
            abc = _cell_text(chunk[2])
            mask = _cell_text(chunk[3])
            if not mask:
                raise ValueError(f"Строка {idx}: не указана маска")
            if mask not in canonical:
                raise ValueError(f"Строка {idx}: неизвестная маска {mask}")
            cap = mask_digit_capacity(mask)
            try:
                category = normalize_import_category(cap, category)
            except ValueError as exc:
                raise ValueError(f"Строка {idx}: {exc}") from exc
            key = (cap, category, abc, mask)
            if key in seen:
                raise ValueError(
                    f"Строка {idx}: повторяется комбинация разрядность/категория/ABC/маска"
                )
            seen.add(key)
            out.append(
                MaskTypeImportRow(
                    digit_capacity=cap,
                    category=category,
                    abc=abc,
                    mask=mask,
                    type_label=_nullable_cell(chunk[4]),
                    premium=_parse_price_cell(chunk[5], row=idx, column="Премиум"),
                    purchase=_parse_price_cell(chunk[6], row=idx, column="Покупка"),
                )
            )
        return out
    finally:
        wb.close()


def _merge_payload(keep: MaskType, drop: MaskType) -> None:
    if not (keep.type_label or "") and (drop.type_label or ""):
        keep.type_label = drop.type_label
    if keep.premium is None and drop.premium is not None:
        keep.premium = drop.premium
    if keep.purchase is None and drop.purchase is not None:
        keep.purchase = drop.purchase


def _backfill_digit_capacity(db: Session, by_key: dict[tuple[str, str, str, str], MaskType]) -> None:
    for row in list(by_key.values()):
        derived = mask_digit_capacity(row.mask)
        if (row.digit_capacity or "") == derived:
            continue
        target = (derived, row.category or "", row.abc or "", row.mask)
        occupied = by_key.get(target)
        if occupied is not None and occupied is not row:
            _merge_payload(occupied, row)
            by_key.pop(row_key(row), None)
            db.delete(row)
            continue
        by_key.pop(row_key(row), None)
        row.digit_capacity = derived
        by_key[target] = row


def _coerce_directory_categories(
    db: Session, by_key: dict[tuple[str, str, str, str], MaskType]
) -> None:
    for row in list(by_key.values()):
        cap = mask_digit_capacity(row.mask)
        current = row.category or ""
        if cap in {"5", "6"} or current == "":
            target_cat = normalize_import_category(cap, current)
        else:
            continue
        if current == target_cat and (row.digit_capacity or "") == cap:
            continue
        target = (cap, target_cat, row.abc or "", row.mask)
        occupied = by_key.get(target)
        old = row_key(row)
        if occupied is not None and occupied is not row:
            _merge_payload(occupied, row)
            by_key.pop(old, None)
            db.delete(row)
            continue
        by_key.pop(old, None)
        row.digit_capacity = cap
        row.category = target_cat
        by_key[target] = row


def ensure_mask_types_seeded(db: Session) -> int:
    rows = list(db.scalars(select(MaskType)).all())
    by_key = {row_key(row): row for row in rows}
    _backfill_digit_capacity(db, by_key)
    _coerce_directory_categories(db, by_key)
    db.flush()
    to_add: list[MaskType] = []
    for mask in all_beauty_masks():
        cap = mask_digit_capacity(mask)
        for category in required_categories(cap):
            key = (cap, category, "", mask)
            if key in by_key:
                continue
            row = MaskType(
                id=uuid4(),
                digit_capacity=cap,
                category=category,
                abc="",
                mask=mask,
            )
            to_add.append(row)
            by_key[key] = row
    if to_add:
        db.add_all(to_add)
    db.commit()
    return len(to_add)


class MaskTypesService:
    TEXT_COLUMNS: dict[str, Any] = {
        "digit_capacity": MaskType.digit_capacity,
        "category": MaskType.category,
        "abc": MaskType.abc,
        "mask": MaskType.mask,
        "type_label": MaskType.type_label,
    }
    PRICE_COLUMNS = {
        "premium": MaskType.premium,
        "purchase": MaskType.purchase,
    }
    FACET_COLUMNS = frozenset({*TEXT_COLUMNS, *PRICE_COLUMNS})
    _ORDER = (
        MaskType.mask.asc(),
        MaskType.digit_capacity.asc(),
        MaskType.category.asc(),
        MaskType.abc.asc(),
    )

    def __init__(self, db: Session):
        self.db = db

    @staticmethod
    def parse_filters(filters: str | dict[str, list[str]] | None) -> dict[str, list[str]]:
        if filters is None or filters == "":
            return {}
        if isinstance(filters, dict):
            data = filters
        else:
            try:
                data = json.loads(filters)
            except json.JSONDecodeError as exc:
                raise ValueError("filters must be a JSON object") from exc
        if not isinstance(data, dict):
            raise ValueError("filters must be a JSON object")
        out: dict[str, list[str]] = {}
        for key, values in data.items():
            if not isinstance(key, str):
                continue
            if isinstance(values, str):
                values = [values]
            if not isinstance(values, list):
                continue
            cleaned = [str(v) for v in values if v is not None]
            if cleaned:
                out[key] = cleaned
        return out

    def _empty_predicate(self, col: ColumnElement[Any]) -> ColumnElement[Any]:
        return or_(col.is_(None), cast(col, String) == "")

    def _apply_mask_q(self, stmt: Select, mask_q: str | None) -> Select:
        if mask_q:
            stmt = stmt.where(MaskType.mask.ilike(f"%{mask_q}%"))
        return stmt

    def _apply_facet_filters(
        self,
        stmt: Select,
        filters: dict[str, list[str]],
        *,
        exclude_column: str | None = None,
    ) -> Select:
        for field, values in filters.items():
            if field == exclude_column:
                continue
            if field not in self.FACET_COLUMNS:
                continue
            if not values:
                continue
            include_empty = EMPTY_TOKEN in values or "" in values
            concrete = [v for v in values if v not in (EMPTY_TOKEN, "")]
            if field in self.PRICE_COLUMNS:
                col = self.PRICE_COLUMNS[field]
                preds = []
                nums = [n for n in (_parse_price_token(v) for v in concrete) if n is not None]
                if nums:
                    preds.append(func.round(col).in_([int(n) for n in nums]))
                if include_empty:
                    preds.append(col.is_(None))
                if preds:
                    stmt = stmt.where(or_(*preds) if len(preds) > 1 else preds[0])
                continue
            col = self.TEXT_COLUMNS[field]
            preds = []
            if concrete:
                preds.append(col.in_(concrete))
            if include_empty:
                preds.append(self._empty_predicate(col))
            if preds:
                stmt = stmt.where(or_(*preds) if len(preds) > 1 else preds[0])
        return stmt

    def _filtered_stmt(
        self,
        *,
        mask_q: str | None = None,
        filters: dict[str, list[str]] | None = None,
        exclude_column: str | None = None,
    ) -> Select:
        stmt = select(MaskType)
        stmt = self._apply_facet_filters(stmt, filters or {}, exclude_column=exclude_column)
        return self._apply_mask_q(stmt, mask_q)

    def list_items(self) -> list[MaskTypeItem]:
        rows = self.db.scalars(select(MaskType).order_by(*self._ORDER)).all()
        return [_item_from_row(row) for row in rows]

    def list_page(
        self,
        *,
        page: int,
        page_size: int,
        mask_q: str | None = None,
        filters: dict[str, list[str]] | None = None,
    ) -> Page[MaskTypeItem]:
        stmt = self._filtered_stmt(mask_q=mask_q, filters=filters)
        count_stmt = select(func.count()).select_from(stmt.order_by(None).subquery())
        total = self.db.scalar(count_stmt) or 0
        stmt = stmt.order_by(*self._ORDER)
        stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        rows = self.db.scalars(stmt).all()
        return Page.of(
            [_item_from_row(row) for row in rows],
            page=page,
            page_size=page_size,
            total=total,
        )

    def _facet_value_expr(self, column: str) -> ColumnElement[Any]:
        if column in self.PRICE_COLUMNS:
            return func.round(self.PRICE_COLUMNS[column])
        col = self.TEXT_COLUMNS.get(column)
        if col is None:
            raise ValueError(f"Unsupported facet column: {column}")
        return col

    def _format_facet_value(self, column: str, raw: Any) -> str:
        if raw is None:
            return ""
        if column in self.PRICE_COLUMNS:
            return _format_price_value(raw)
        return str(raw)

    def list_facets(
        self,
        *,
        column: str,
        filters: dict[str, list[str]] | None = None,
        mask_q: str | None = None,
        q: str | None = None,
        limit: int = 200,
        offset: int = 0,
    ) -> FacetResponse:
        if column not in self.FACET_COLUMNS:
            raise ValueError(f"Unsupported facet column: {column}")

        value_expr = self._facet_value_expr(column)
        if column in self.TEXT_COLUMNS:
            value_expr = func.nullif(func.btrim(cast(value_expr, String)), "")

        stmt = self._filtered_stmt(
            mask_q=mask_q,
            filters=filters,
            exclude_column=column,
        )
        base = (
            stmt.with_only_columns(
                value_expr.label("facet_value"),
                MaskType.id.label("row_id"),
            )
            .order_by(None)
            .subquery()
        )
        grouped = select(
            base.c.facet_value,
            func.count().label("cnt"),
        ).group_by(base.c.facet_value)
        if q:
            grouped = grouped.where(cast(base.c.facet_value, String).ilike(f"%{q}%"))
        grouped = grouped.order_by(
            func.count().desc(),
            cast(base.c.facet_value, String).asc().nulls_last(),
        )
        rows = self.db.execute(grouped.offset(offset).limit(limit + 1)).all()
        truncated = len(rows) > limit
        rows = rows[:limit]
        items = [
            FacetItem(value=self._format_facet_value(column, raw), count=int(cnt))
            for raw, cnt in rows
        ]
        return FacetResponse(column=column, items=items, truncated=truncated)

    def write_xlsx(self, path: str | Path) -> int:
        items = self.list_items()
        wb = open_styled_workbook(str(path), constant_memory=True)
        try:
            ws = wb.add_worksheet(MASK_TYPES_XLSX_SHEET)
            writer = StyledSheetWriter(wb, ws, MASK_TYPES_XLSX_HEADERS)
            for item in items:
                writer.write_row(
                    [
                        item.digit_capacity,
                        item.category,
                        item.abc,
                        item.mask,
                        item.type_label or "",
                        _xlsx_price(item.premium),
                        _xlsx_price(item.purchase),
                    ],
                    fill_color=mask_row_fill_color(item.category, item.premium),
                )
            writer.finalize()
        finally:
            wb.close()
        return len(items)

    def upsert_from_xlsx(self, data: bytes) -> MaskTypesLoadResult:
        ensure_mask_types_seeded(self.db)
        parsed = parse_mask_types_xlsx(data)
        existing_rows = list(self.db.scalars(select(MaskType)).all())
        by_key = {row_key(row): row for row in existing_rows}
        by_mask: dict[str, list[MaskType]] = {}
        for row in existing_rows:
            by_mask.setdefault(row.mask, []).append(row)

        inserted = 0
        updated = 0
        deleted = 0

        def forget(row: MaskType) -> None:
            by_key.pop(row_key(row), None)
            bucket = by_mask.get(row.mask)
            if bucket:
                by_mask[row.mask] = [item for item in bucket if item is not row]

        def add_row(item: MaskTypeImportRow) -> MaskType:
            row = MaskType(
                id=uuid4(),
                digit_capacity=item.digit_capacity,
                category=item.category,
                abc=item.abc,
                mask=item.mask,
                type_label=item.type_label,
                premium=item.premium,
                purchase=item.purchase,
            )
            self.db.add(row)
            self.db.flush()
            by_key[row_key(row)] = row
            by_mask.setdefault(row.mask, []).append(row)
            return row

        by_file: dict[str, list[MaskTypeImportRow]] = {}
        for item in parsed:
            by_file.setdefault(item.mask, []).append(item)

        for mask, file_items in by_file.items():
            for item in file_items:
                key = (item.digit_capacity, item.category, item.abc, item.mask)
                exact = by_key.get(key)
                if exact is not None:
                    apply_payload(exact, item)
                    updated += 1
                    continue
                add_row(item)
                inserted += 1

            desired = {(it.category, it.abc) for it in file_items}
            cap = mask_digit_capacity(mask)
            for row in list(by_mask.get(mask, [])):
                combo = (row.category or "", row.abc or "")
                if combo in desired:
                    continue
                if is_required_key(
                    digit_capacity=row.digit_capacity or cap,
                    category=row.category or "",
                    abc=row.abc or "",
                ):
                    continue
                forget(row)
                self.db.delete(row)
                deleted += 1

        self.db.commit()
        return MaskTypesLoadResult(
            ok=True,
            count=len(parsed),
            updated=updated,
            inserted=inserted,
            message=(
                f"Обновлено: {updated}, добавлено: {inserted}, удалено: {deleted}"
            ),
        )
