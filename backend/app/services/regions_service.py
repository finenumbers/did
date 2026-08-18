from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.regions_directory import RegionsDirectory
from app.providers.msisdn_split import DIGIT_CAPACITY_MAX, DIGIT_CAPACITY_MIN
from app.schemas.regions import RegionCityItem, RegionsLoadResult
from app.services.xlsx_style import StyledSheetWriter, open_styled_workbook

REGIONS_XLSX_HEADERS = ("ABC", "Разрядность", "Город", "Регион")
REGIONS_XLSX_SHEET = "Регионы"
MAX_IMPORT_BYTES = 5 * 1024 * 1024


@dataclass(frozen=True)
class RegionImportRow:
    abc: str
    digit_capacity: int
    city_name: str
    region_name: str | None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            return str(value).strip()
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def _abc_from_cell(value: object) -> str:
    text = _cell_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _capacity_from_cell(value: object) -> int:
    if isinstance(value, bool):
        raise ValueError("некорректная разрядность")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError("некорректная разрядность")
        return int(value)
    text = _cell_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if not text or not text.isdigit():
        raise ValueError("некорректная разрядность")
    return int(text)


def _row_empty(values: list[object]) -> bool:
    return all(_cell_text(v) == "" for v in values)


def parse_regions_xlsx(data: bytes) -> list[RegionImportRow]:
    """Parse first sheet. Raises ValueError; does not touch the database."""
    if not data:
        raise ValueError("Пустой файл")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("Файл слишком большой")
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось прочитать XLSX") from exc
    try:
        if not wb.worksheets:
            raise ValueError("В файле нет листов")
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ValueError("Нет заголовков") from exc
        got = tuple(_cell_text(c) for c in list(header or ())[:4])
        if got != REGIONS_XLSX_HEADERS:
            raise ValueError("Заголовки должны быть: ABC, Разрядность, Город, Регион")
        out: list[RegionImportRow] = []
        seen: set[str] = set()
        for idx, raw in enumerate(rows_iter, start=2):
            cells = list(raw or ())
            while len(cells) < 4:
                cells.append(None)
            chunk = cells[:4]
            if _row_empty(chunk):
                continue
            abc = _abc_from_cell(chunk[0])
            if not abc.isdigit():
                raise ValueError(f"Строка {idx}: некорректный ABC")
            try:
                cap = _capacity_from_cell(chunk[1])
            except ValueError as exc:
                raise ValueError(f"Строка {idx}: некорректная разрядность") from exc
            if cap < DIGIT_CAPACITY_MIN or cap > DIGIT_CAPACITY_MAX:
                raise ValueError(f"Строка {idx}: разрядность должна быть от 5 до 7")
            if len(abc) + cap != 10:
                raise ValueError(
                    f"Строка {idx}: длина ABC плюс разрядность должны давать 10"
                )
            city = _cell_text(chunk[2])
            if not city:
                raise ValueError(f"Строка {idx}: не указан город")
            region = _cell_text(chunk[3]) or None
            if abc in seen:
                raise ValueError(f"Строка {idx}: повторяется ABC {abc}")
            seen.add(abc)
            out.append(
                RegionImportRow(
                    abc=abc,
                    digit_capacity=cap,
                    city_name=city,
                    region_name=region,
                )
            )
        return out
    finally:
        wb.close()


class RegionsService:
    """Regions page table. Snapshot only via XLSX import/export."""

    def __init__(self, db: Session):
        self.db = db

    def list_cities(self) -> list[RegionCityItem]:
        rows = self.db.scalars(
            select(RegionsDirectory).order_by(
                RegionsDirectory.abc.asc(),
                RegionsDirectory.city_name.asc().nulls_last(),
            )
        ).all()
        return [
            RegionCityItem(
                id=row.id,
                abc=row.abc,
                digit_capacity=int(row.digit_capacity),
                city_name=row.city_name,
                region_name=row.region_name,
            )
            for row in rows
        ]

    def write_xlsx(self, path: str | Path) -> int:
        items = self.list_cities()
        wb = open_styled_workbook(str(path), constant_memory=True)
        try:
            ws = wb.add_worksheet(REGIONS_XLSX_SHEET)
            writer = StyledSheetWriter(wb, ws, REGIONS_XLSX_HEADERS)
            for item in items:
                writer.write_row(
                    [
                        item.abc,
                        item.digit_capacity,
                        item.city_name,
                        item.region_name or "",
                    ]
                )
            writer.finalize()
        finally:
            wb.close()
        return len(items)

    def replace_from_xlsx(self, data: bytes) -> RegionsLoadResult:
        parsed = parse_regions_xlsx(data)
        loaded_at = datetime.now(timezone.utc)
        self.db.execute(delete(RegionsDirectory))
        if parsed:
            self.db.add_all(
                [
                    RegionsDirectory(
                        id=uuid4(),
                        abc=row.abc,
                        digit_capacity=row.digit_capacity,
                        city_name=row.city_name,
                        region_name=row.region_name,
                        loaded_at=loaded_at,
                    )
                    for row in parsed
                ]
            )
        self.db.commit()
        return RegionsLoadResult(
            ok=True,
            count=len(parsed),
            message=f"Загружено строк: {len(parsed)}",
        )
